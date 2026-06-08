"""
Module: excel.py
Description: Utility functions for exporting schedules to Excel files. This module provides functionality to take the generated schedule from the Scheduler class and create a visually organized Excel file that shows daily assignments for each doctor, while also marking any days they are unavailable. This allows for easy sharing and review of the schedule in a familiar format. The Excel export includes features such as:
- A header row indicating the day of the week and date.
"""

from scheduler import ShiftScheduler
import openpyxl
from openpyxl.styles import PatternFill


class ExcelUtils:
    """Utility class for exporting schedules to Excel files."""

    def export_schedule_to_excel(self, filename: str, scheduler: ShiftScheduler):
        """
        Exports the generated schedule to an Excel file. 
        showing daily assignments and marking doctor unavailability.
        """
        schedule = {}
        for (day_idx, _, shift, doc), var in scheduler.shift_assignments.items():
            if scheduler.solver.value(var) == 1:
                schedule[(day_idx, doc)] = shift.name

        wb = openpyxl.Workbook()
        ws = wb.active

        unavailable_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid")

        # Row 1: day of week
        for day_idx, date in enumerate(scheduler.dates):
            ws.cell(row=1, column=day_idx + 2, value=date.strftime('%a'))

        # Row 2: day number
        for day_idx, date in enumerate(scheduler.dates):
            ws.cell(row=2, column=day_idx + 2, value=date.day)

        for row_idx, doctor in enumerate(scheduler.department.doctor_order):
            ws.cell(row=row_idx+3, column=1, value=doctor.name)
            print(f"Doctor: {doctor.name}")
            for day_idx, _ in enumerate(scheduler.dates):
                cell = ws.cell(row=row_idx + 3, column=day_idx + 2)
                if scheduler.dates[day_idx] in doctor.unavailability:
                    cell.fill = unavailable_fill
                else:
                    cell.value = schedule.get((day_idx, doctor), "")

        for row_idx, doctor in enumerate(scheduler.department.doctor_order):
            for day_idx, _ in enumerate(scheduler.dates):
                cell_value = schedule.get((day_idx, doctor), "")
                ws.cell(row=row_idx+3, column=day_idx+2, value=cell_value)

            # weekend_fill = PatternFill(
            #     start_color="000000", end_color="000000", fill_type="solid")
            # weekend_font = Font(color="FFFFFF", bold=True)
        wb.save(filename)
