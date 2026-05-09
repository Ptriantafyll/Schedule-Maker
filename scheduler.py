"""Shift scheduling engine using Google OR-Tools CP-SAT solver.

Builds and solves a constraint programming model to generate
optimal monthly duty schedules for a hospital department.
"""

import datetime
import calendar
import math
import openpyxl
from ortools.sat.python import cp_model
from openpyxl.styles import PatternFill
from models import Department, Doctor, Position, Shift, Team


class ShiftScheduler:  # pylint: disable=too-many-instance-attributes
    """Shift scheduling engine using Google OR-Tools CP-SAT solver."""

    def __init__(self, department: Department):
        self.department = department
        self.is_weekend = {}
        self.model = cp_model.CpModel()
        self.shift_assignments = {}
        self.penalties = []
        self.rewards = []
        self.solver = cp_model.CpSolver()
        self.dates = []
        self.weekend_off_vars = {}

    def _calculate_days_for_schedule(self, month: int, year: int) -> list[datetime.date]:
        """Returns the list of dates and a weekend lookup dict for the given month."""

        first_day = datetime.date(year, month, 1)
        days_in_month = calendar.monthrange(year, month)[1]
        self.dates = [first_day + datetime.timedelta(days=i)
                      for i in range(days_in_month)]

        self.is_weekend = {d: (d.weekday() >= 5) for d in self.dates}

    def _get_assignment_vars_for(self, day_index=None, position=None, shift=None, doctor=None):
        """Returns all shift assignment variables matching the given filters."""
        return [
            var for (d, p, s, doc), var in self.shift_assignments.items()
            if (day_index is None or d == day_index)
            and (position is None or p == position)
            and (shift is None or s == shift)
            and (doctor is None or doc == doctor)
        ]

    def _get_weekends(self) -> list[list[int]]:
        """Returns list of (fri, sat, sun) day index tuples for complete weekends."""
        weekends = []

        for day_index, date in enumerate(self.dates):
            if date.weekday() == 4 and day_index + 2 < len(self.dates):
                weekends.append([day_index, day_index+1, day_index+2])

        return weekends

    def _build_model(self):
        """Creates the CP-SAT model and shift assignment variables."""

        for position in self.department.positions:
            position_doctors = position.eligible_doctors if position.eligible_doctors else self.department.doctors

            for doctor in position_doctors:
                self._create_assignment_variable(position, doctor)

    def _create_assignment_variable(self, position: Position, doctor: Doctor):
        """Creates shift assignment variables for a doctor in a position."""
        pre_assignments = set(doctor.pre_assignments)

        for shift in position.shifts:
            for day_index, date in enumerate(self.dates):
                if date in doctor.unavailability:
                    continue
                if date.weekday() not in position.duty_days:
                    continue

                var = self.model.new_bool_var(
                    f"shift_assignment_{day_index}_{position}_{shift}_{doctor}")
                self.shift_assignments[(
                    day_index, position, shift, doctor)] = var

                if (date, shift) in pre_assignments:
                    self.model.add(self.shift_assignments[(
                        day_index, position, shift, doctor)] == 1)
                else:
                    if pre_assignments:
                        self.model.add(self.shift_assignments[(
                            day_index, position, shift, doctor)] == 0)

    def _calculate_duties_per_doctor(self):
        """Returns the expected duties per non-pre-assigned doctor."""
        pre_assigned_duties = sum(len(doctor.pre_assignments)
                                  for doctor in self.department.doctors if doctor.pre_assignments)
        total_duties = sum(
            shift.doctors_per_shift *
            sum(1 for d in self.dates if d.weekday() in position.duty_days)
            for position in self.department.positions
            for shift in position.shifts
        )
        duties_needed = total_duties - pre_assigned_duties
        doctors_available = sum(
            1 for doctor in self.department.doctors if not doctor.pre_assignments)
        return duties_needed // doctors_available

    def _combine_objectives(self):
        self.model.minimize(sum(self.penalties) - sum(self.rewards))

    def _add_hard_constraint_one_shift_per_doctor_per_day(self):
        """Ensures a doctor can only be assigned to at most 1 shift per day."""
        for doctor in self.department.doctors:
            for day_index in range(len(self.dates)):
                day_vars = self._get_assignment_vars_for(
                    day_index=day_index, doctor=doctor)
                if day_vars:
                    self.model.add(sum(day_vars) <= 1)

    def _add_hard_constraint_doctors_per_shift(self):
        """Adds a hard constraint that a shift must have the exact number of doctor as specified"""

        for day_index, date in enumerate(self.dates):
            for position in self.department.positions:
                if date.weekday() not in position.duty_days:
                    continue

                for shift in position.shifts:

                    self.model.add(
                        sum(self._get_assignment_vars_for(
                            day_index, position, shift))
                        == shift.doctors_per_shift
                    )

    def _add_hard_constraint_no_consecutive_shifts(self):
        """Adds a hard constraint that a doctor cannot be on duty for 2 consecutive days"""

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue

            for day_index in range(len(self.dates) - 1):
                today_shifts = self._get_assignment_vars_for(
                    day_index=day_index, doctor=doctor)

                tomorrow_shifts = self._get_assignment_vars_for(
                    day_index=day_index + 1, doctor=doctor)

                self.model.add(sum(today_shifts + tomorrow_shifts) <= 1)

    def _add_hard_constraint_max_duties_per_doc_per_month(self):
        """Adds a hard constraint that sets the max duties per month a doctor can do"""

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue

            shifts = self._get_assignment_vars_for(doctor=doctor)
            self.model.add(
                sum(shifts) <= self.department.config.max_duties_per_month
            )

    def _add_hard_constraint_one_full_weekend_off_per_doctor(self):
        """Ensures every doctor has at least one weekend (Sat+Sun) off."""
        weekends = self._get_weekends()

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue
            weekend_off_vars = []

            for weekend_index, (_, sat, sun) in enumerate(weekends):
                weekend_off = self.model.new_bool_var(
                    f"full_weekend_off_{weekend_index}_{doctor}"
                )

                psk_shifts = (
                    self._get_assignment_vars_for(
                        day_index=sat, doctor=doctor
                    ) + self._get_assignment_vars_for(
                        day_index=sun, doctor=doctor
                    )
                )

                if not psk_shifts:
                    self.model.add(weekend_off == 1)
                else:
                    self.model.add(sum(psk_shifts) + len(psk_shifts)
                                   * weekend_off <= len(psk_shifts))
                weekend_off_vars.append(weekend_off)

            self.weekend_off_vars[doctor] = weekend_off_vars

            self.model.add(sum(weekend_off_vars) >= 1)

    def _add_hard_constraint_balanced_total_duties_across_doctors(self):
        """Ensures all doctors have the same number of duties (+-1)"""
        min_duties = self._calculate_duties_per_doctor()
        max_duties = min_duties + 1

        # min_duties = 5
        # max_duties = 6

        print(f"Balance: min={min_duties}, max={max_duties}")

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue
            available = len(self._get_assignment_vars_for(doctor=doctor))
            if available < 4:
                print(f"  {doctor.name}: only {available} assignment vars")

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue

            doctor_assignments = self._get_assignment_vars_for(doctor=doctor)
            self.model.add(sum(doctor_assignments) <= max_duties)
            self.model.add(sum(doctor_assignments) >= min_duties)

    def _add_hard_constraint_balanced_weekend_duties_across_doctors(self):
        """Ensures all doctors have the same number of weekend duties (+-1)"""
        pre_assigned_weekends = 0

        for doctor in self.department.doctors:
            if not doctor.pre_assignments:
                continue

            for (pre_assigned_date, _) in doctor.pre_assignments:
                if self.is_weekend[pre_assigned_date]:
                    pre_assigned_weekends += 1

        total_weekend_duties = sum(
            shift.doctors_per_shift *
            sum(1 for d in self.dates if
                self.is_weekend[d] and d.weekday() in position.duty_days)
            for position in self.department.positions
            for shift in position.shifts
        )

        weekend_duties_needed = total_weekend_duties - pre_assigned_weekends
        doctors_available = sum(
            1 for doctor in self.department.doctors if not doctor.pre_assignments)

        min_weekend_duties = weekend_duties_needed // doctors_available
        max_weekend_duties = min_weekend_duties + 1

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue
            assigned_weekends = []

            for day_index, date in enumerate(self.dates):
                if not self.is_weekend[date]:
                    continue

                assigned_weekends.extend(self._get_assignment_vars_for(
                    doctor=doctor, day_index=day_index))

            self.model.add(sum(assigned_weekends) <= max_weekend_duties)
            self.model.add(sum(assigned_weekends) >= min_weekend_duties)

    def _get_day_off_vars_for_team(self, team, day_index):
        """Returns shift vars from the previous day that grant a day off for a team."""
        day_off_vars = []
        for doctor in team.doctors:
            for position in self.department.positions:
                for shift in position.shifts:
                    if not shift.grants_day_off:
                        continue

                    key = (day_index, position, shift, doctor)
                    if key in self.shift_assignments:
                        day_off_vars.append(self.shift_assignments[key])

        return day_off_vars

    def _add_hard_constraint_max_one_team_day_off(self):
        """Ensures at most 1 doctor per team has a post-shift day off on the same day."""
        for day_index in range(1, len(self.dates)):
            for team in self.department.teams:
                yesterday_day_off_shifts = self._get_day_off_vars_for_team(
                    team, day_index-1)
            self.model.add(sum(yesterday_day_off_shifts) <= 1)

    def _add_soft_constraint_penalize_every_other_day_on_duty(self):
        """Penalizes on-off-on patterns where a doctor works day N and day N+2."""
        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue

            for day_index in range(len(self.dates) - 2):
                day_a_var = self._get_assignment_vars_for(
                    day_index=day_index, doctor=doctor)
                day_c_var = self._get_assignment_vars_for(
                    day_index=day_index+2, doctor=doctor)

                if (not day_a_var) or (not day_c_var):
                    continue

                is_every_other = self.model.new_bool_var(
                    f"every_other_penalty_{day_index}_{doctor}_"
                )

                self.model.add(sum(day_a_var) + sum(day_c_var) ==
                               2).only_enforce_if(is_every_other)
                self.model.add(sum(day_a_var) + sum(day_c_var) !=
                               2).only_enforce_if(is_every_other.Not())

                self.penalties.append(
                    self.department.config.w_every_other_penalty * is_every_other)

    def _add_soft_constraint_penalize_duty_gap(self, gap_size: int, weight: int):
        """
        Penalizes short gaps between duties, 
        e.g. working on day N and then again on day N+2 with only one day off in between.
        """
        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue

            for day_index in range(len(self.dates) - gap_size):
                day_a_var = self._get_assignment_vars_for(
                    day_index=day_index, doctor=doctor)
                day_c_var = self._get_assignment_vars_for(
                    day_index=day_index+gap_size, doctor=doctor)

                if (not day_a_var) or (not day_c_var):
                    continue

                is_short_gap = self.model.new_bool_var(
                    f"duty_gap_{day_index}_{doctor}_"
                )

                self.model.add(sum(day_a_var) + sum(day_c_var) ==
                               2).only_enforce_if(is_short_gap)
                self.model.add(sum(day_a_var) + sum(day_c_var) !=
                               2).only_enforce_if(is_short_gap.Not())

                self.penalties.append(weight * is_short_gap)

    # Wrappers for tests
    def _add_soft_constraint_penalize_every_other_day_on_duty(self):
        self._add_soft_constraint_penalize_duty_gap(
            2, self.department.config.w_every_other_penalty)

    def _add_soft_constraint_penalize_short_gaps_between_duties(self):
        self._add_soft_constraint_penalize_duty_gap(
            3, self.department.config.w_gap_penalty)

    def _add_soft_constraint_spread_duties_across_month(self):
        """Penalizes uneven distribution of duties across month blocks per doctor."""
        duties_per_doctor = self._calculate_duties_per_doctor()
        num_blocks = self.department.config.month_blocks
        block_size = math.ceil(len(self.dates) / num_blocks)
        ideal = duties_per_doctor / num_blocks

        for block in range(num_blocks):
            block_start = block * block_size
            block_end = min((block + 1) * block_size, len(self.dates))

            for doctor in self.department.doctors:
                if doctor.pre_assignments:
                    continue

                block_duties = sum(
                    sum(self._get_assignment_vars_for(
                        day_index=d, doctor=doctor))
                    for d in range(block_start, block_end)
                )

                deviation = self.model.new_int_var(
                    0, block_size, f"block_{block}_deviation_for_{doctor}")
                self.model.add(deviation >= block_duties -
                               int(math.ceil(ideal)))
                self.model.add(deviation >= int(
                    math.floor(ideal)) - block_duties)

                self.penalties.append(
                    self.department.config.w_block_dev_penalty * deviation)

    def _add_soft_constraint_reward_full_weekends_off(self):
        """Rewards doctors for having full weekends (Fri+Sat+Sun) off."""
        weekends = self._get_weekends()

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue

            for weekend_index, (fri, sat, sun) in enumerate(weekends):
                weekend_off = self.model.new_bool_var(
                    f"full_weekend_off_{weekend_index}_{doctor}")

                psk_shifts = (
                    self._get_assignment_vars_for(day_index=fri, doctor=doctor)
                    + self._get_assignment_vars_for(day_index=sat, doctor=doctor)
                    + self._get_assignment_vars_for(day_index=sun, doctor=doctor)
                )

                if not psk_shifts:
                    self.model.add(weekend_off == 1)
                else:
                    self.model.add(sum(psk_shifts) + len(psk_shifts)
                                   * weekend_off <= len(psk_shifts))

                self.rewards.append(
                    self.department.config.w_full_wkend_off_bonus * weekend_off)

    def _add_soft_constraint_balance_full_weekends_off(self):
        """Penalizes if the number of full weekends off is not balanced across doctors."""
        weekends = self._get_weekends()
        doctor_weekend_counts = []

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue
            vars_for_doctor = self.weekend_off_vars.get(doctor, [])
            if not vars_for_doctor:
                continue
            count_var = self.model.new_int_var(
                0, len(weekends), f"weekend_off_count_{doctor}")
            self.model.add(count_var == sum(vars_for_doctor))
            doctor_weekend_counts.append(count_var)

        if len(doctor_weekend_counts) < 2:
            return

        min_var = self.model.new_int_var(0, len(weekends), "min_weekends_off")
        max_var = self.model.new_int_var(0, len(weekends), "max_weekends_off")
        self.model.add_min_equality(min_var, doctor_weekend_counts)
        self.model.add_max_equality(max_var, doctor_weekend_counts)

        self.penalties.append(
            self.department.config.w_balance_full_wkends_off * (max_var - min_var))

    def _add_soft_constraint_balance_saturday_sunday_duties(self):
        """Penalizes if a doctor has unbalanced Saturday vs Sunday duties."""

        for doctor in self.department.doctors:
            if doctor.pre_assignments:
                continue

            saturday_duties = sum(
                vars
                for day in range(len(self.dates))
                if self.dates[day].weekday() == 5
                for vars in self._get_assignment_vars_for(day_index=day, doctor=doctor)
            )

            sunday_duties = sum(
                vars
                for day in range(len(self.dates))
                if self.dates[day].weekday() == 6
                for vars in self._get_assignment_vars_for(day_index=day, doctor=doctor)
            )

            saturdays_of_month = sum(
                1 for d in self.dates if d.weekday() == 5)
            sundays_of_month = sum(
                1 for d in self.dates if d.weekday() == 6)

            sat_sun_deviation = self.model.new_int_var(
                0, max(saturdays_of_month, sundays_of_month), f"sat_sun_deviation_{doctor}")

            self.model.add(sat_sun_deviation >=
                           saturday_duties - sunday_duties)
            self.model.add(sat_sun_deviation >=
                           sunday_duties - saturday_duties)

            self.penalties.append(
                self.department.config.w_diff_wkend_duty_day * sat_sun_deviation)

    def _debug_print_capacity(self):
        """Prints debug info about doctor availability per position."""
        for position in self.department.positions:
            eligible = len(position.eligible_doctors) if position.eligible_doctors else len(
                self.department.doctors)
            duty_day_count = sum(
                1 for d in self.dates if d.weekday() in position.duty_days)
            needed_per_day = sum(s.doctors_per_shift for s in position.shifts)
            total_needed = duty_day_count * needed_per_day
            print(f"{position.name}: {eligible} eligible doctors, {duty_day_count} duty days, {total_needed} total assignments needed")

    def create_schedule(self, month: int, year: int):
        """
        Main method to create the schedule for the given month and year.
        Builds the model, adds constraints, and solves it.
        """
        self._calculate_days_for_schedule(month=month, year=year)
        self._build_model()
        self._debug_print_capacity()

        # Check if no-consecutive is feasible per position
        for position in self.department.positions:
            eligible = position.eligible_doctors if position.eligible_doctors else self.department.doctors
            duty_day_count = sum(
                1 for d in self.dates if d.weekday() in position.duty_days)
            needed_per_day = sum(s.doctors_per_shift for s in position.shifts)
            # With no-consecutive, a doctor can work at most ceil(duty_days/2) days
            max_per_doctor = (duty_day_count + 1) // 2
            available_slots = sum(
                max_per_doctor for d in eligible if not d.pre_assignments)
            print(
                f"{position.name}: need {duty_day_count * needed_per_day}, max available ~{available_slots}")

        for _, date in enumerate(self.dates):
            total = 0
            for position in self.department.positions:
                if date.weekday() not in position.duty_days:
                    continue
                total += sum(s.doctors_per_shift for s in position.shifts)
            if total > 0:
                print(f"{date} ({date.strftime('%a')}): {total} doctors needed")

        self._add_hard_constraint_doctors_per_shift()
        self._add_hard_constraint_one_shift_per_doctor_per_day()
        self._add_hard_constraint_no_consecutive_shifts()
        self._add_hard_constraint_max_duties_per_doc_per_month()
        self._add_hard_constraint_balanced_total_duties_across_doctors()
        self._add_hard_constraint_balanced_weekend_duties_across_doctors()
        self._add_hard_constraint_max_one_team_day_off()
        self._add_hard_constraint_one_full_weekend_off_per_doctor()
        self._add_soft_constraint_penalize_every_other_day_on_duty()
        self._add_soft_constraint_spread_duties_across_month()
        self._add_soft_constraint_reward_full_weekends_off()
        self._add_soft_constraint_balance_full_weekends_off()
        self._add_soft_constraint_balance_saturday_sunday_duties()
        self._combine_objectives()

        status = self.solver.Solve(self.model)

        print(f"Solver status: {status}")
        print(f"Status name: {self.solver.status_name(status)}")

        return status

    def _print_daily_assignments(self):
        """Prints the daily schedule with assignments."""
        for day_index, date in enumerate(self.dates):
            day_total = 0
            assignments = []
            for (day_idx, pos, shift, doc), var in self.shift_assignments.items():
                if day_idx == day_index and self.solver.value(var) == 1:
                    assignments.append(
                        f"    {pos.name} / {shift.name}: {doc.name}")
                    day_total += 1
            print(f"{date} ({date.strftime('%a')}) — {day_total} assigned")
            for a in assignments:
                print(a)

    def _print_doctor_workloads(self):
        """Prints total and weekend duties per doctor."""
        for doctor in self.department.doctors:
            total = sum(
                self.solver.Value(var)
                for var in self._get_assignment_vars_for(doctor=doctor)
            )
            weekend_total = sum(
                self.solver.Value(var)
                for (d, p, s, doc), var in self.shift_assignments.items()
                if doc == doctor and self.is_weekend[self.dates[d]]
            )
            if total > 0:
                print(f"  {doctor.name}: {total} (weekend: {weekend_total})")

    def print_schedule(self):
        """Prints the generated schedule."""
        self._print_daily_assignments()
        print("\n--- Duties per doctor ---")
        self._print_doctor_workloads()

    def export_to_exel(self, filename: str):
        """
        Exports the generated schedule to an Excel file. 
        showing daily assignments and marking doctor unavailability.
        """
        schedule = {}
        for (day_idx, _, shift, doc), var in self.shift_assignments.items():
            if self.solver.value(var) == 1:
                schedule[(day_idx, doc)] = shift.name

        wb = openpyxl.Workbook()
        ws = wb.active

        unavailable_fill = PatternFill(
            start_color="000000", end_color="000000", fill_type="solid")

        # Row 1: day of week
        for day_idx, date in enumerate(self.dates):
            ws.cell(row=1, column=day_idx + 2, value=date.strftime('%a'))

        # Row 2: day number
        for day_idx, date in enumerate(self.dates):
            ws.cell(row=2, column=day_idx + 2, value=date.day)

        for row_idx, doctor in enumerate(self.department.doctor_order):
            ws.cell(row=row_idx+3, column=1, value=doctor.name)
            print(f"Doctor: {doctor.name}")
            for day_idx, _ in enumerate(self.dates):
                cell = ws.cell(row=row_idx + 3, column=day_idx + 2)
                if self.dates[day_idx] in doctor.unavailability:
                    cell.fill = unavailable_fill
                else:
                    cell.value = schedule.get((day_idx, doctor), "")

        for row_idx, doctor in enumerate(self.department.doctor_order):
            for day_idx, _ in enumerate(self.dates):
                cell_value = schedule.get((day_idx, doctor), "")
                ws.cell(row=row_idx+3, column=day_idx+2, value=cell_value)

            # weekend_fill = PatternFill(
            #     start_color="000000", end_color="000000", fill_type="solid")
            # weekend_font = Font(color="FFFFFF", bold=True)
        wb.save(filename)


if __name__ == "__main__":
    polyzou_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 7),
        datetime.date(2026, 3, 8),
        datetime.date(2026, 3, 9),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
    ]
    eleftherakis_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
        datetime.date(2026, 3, 16),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 27),
    ]
    michailidis_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
        datetime.date(2026, 3, 16),
        datetime.date(2026, 3, 17),
        datetime.date(2026, 3, 18),
        datetime.date(2026, 3, 19),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
        datetime.date(2026, 3, 23),
        datetime.date(2026, 3, 24),
        datetime.date(2026, 3, 25),
        datetime.date(2026, 3, 26),
        datetime.date(2026, 3, 27),
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 29),
        datetime.date(2026, 3, 30),
        datetime.date(2026, 3, 31),

    ]
    petropoulou_unavailability = [
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
    ]
    kassara_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 5),
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 7),
        datetime.date(2026, 3, 8),
        datetime.date(2026, 3, 18),
    ]
    marlafeka_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
        datetime.date(2026, 3, 25),
        datetime.date(2026, 3, 26),
        datetime.date(2026, 3, 27),
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 29),
        datetime.date(2026, 3, 30),
    ]
    efthymiou_unavailability = marlafeka_unavailability
    triantafyllopoulos_unavailability = [
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 29),
    ]
    papakamenos_unavailability = [
        datetime.date(2026, 3, 8),
        datetime.date(2026, 3, 15),
        datetime.date(2026, 3, 22),
        datetime.date(2026, 3, 25),
        datetime.date(2026, 3, 27),
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 29),
        datetime.date(2026, 3, 30),
        datetime.date(2026, 3, 31),
    ]
    karydis_unavailability = [
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
        datetime.date(2026, 3, 25),
        datetime.date(2026, 3, 26),
        datetime.date(2026, 3, 27),
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 29),
    ]
    aravantinos_unavailability = []
    argyropoulos_unavailability = [
        datetime.date(2026, 3, 5),
    ]
    xiotis_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 5),
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 7),
        datetime.date(2026, 3, 8),
        datetime.date(2026, 3, 9),
        datetime.date(2026, 3, 18),
    ]
    nikas_unavailability = [
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 7),
        datetime.date(2026, 3, 8),
        datetime.date(2026, 3, 27),
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 29),
    ]
    adamopoulou_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 22),
        datetime.date(2026, 3, 23),
        datetime.date(2026, 3, 24),
        datetime.date(2026, 3, 25),
        datetime.date(2026, 3, 26),
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 29),
    ]
    mpakalarou_unavailability = [
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
    ]
    papageorgiou_dsp_unavailability = [
        datetime.date(2026, 3, 18),
        datetime.date(2026, 3, 19),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
    ]
    rousia_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 4),
        datetime.date(2026, 3, 5),
        datetime.date(2026, 3, 20),
    ]
    pitsi_unavailability = [
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
        datetime.date(2026, 3, 25),
    ]
    moutsi_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 3),
        datetime.date(2026, 3, 5),
        datetime.date(2026, 3, 8),
        datetime.date(2026, 3, 10),
        datetime.date(2026, 3, 12),
        datetime.date(2026, 3, 15),
        datetime.date(2026, 3, 17),
        datetime.date(2026, 3, 19),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
        datetime.date(2026, 3, 23),
        datetime.date(2026, 3, 24),
        datetime.date(2026, 3, 25),
        datetime.date(2026, 3, 26),
        datetime.date(2026, 3, 29),
        datetime.date(2026, 3, 31),
    ]
    tzikopoulos_unavailability = []
    giannaki_unavailability = [
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
    ]
    gkolemi_unavailability = [
        datetime.date(2026, 3, 5),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
    ]
    tasiopoulos_unavailability = []
    florou_unavailability = []
    spiliotopoulos_unavailability = []
    xrysanthakopoulos_unavailability = [
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
        datetime.date(2026, 3, 25),
        datetime.date(2026, 3, 30),
    ]
    pandi_unavailability = []
    argyriadi_unavailability = [
        datetime.date(2026, 3, 1),
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 3),
        datetime.date(2026, 3, 4),
        datetime.date(2026, 3, 5),
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 7),
        datetime.date(2026, 3, 8),
        datetime.date(2026, 3, 9),
        datetime.date(2026, 3, 10),
        datetime.date(2026, 3, 11),
        datetime.date(2026, 3, 12),
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
    ]
    armeni_unavailability = [
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 9),
        datetime.date(2026, 3, 16),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 23),
        datetime.date(2026, 3, 27),
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 30),
    ]
    papadopoulos_unavailability = [
        datetime.date(2026, 3, 2),
        datetime.date(2026, 3, 9),
        datetime.date(2026, 3, 10),
        datetime.date(2026, 3, 11),
        datetime.date(2026, 3, 12),
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
        datetime.date(2026, 3, 16),
        datetime.date(2026, 3, 25),
    ]
    erotokritou_unavailability = [
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 14),
        datetime.date(2026, 3, 15),
        datetime.date(2026, 3, 16),
        datetime.date(2026, 3, 17),
        datetime.date(2026, 3, 18),
        datetime.date(2026, 3, 19),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
        datetime.date(2026, 3, 23),
        datetime.date(2026, 3, 24),
        datetime.date(2026, 3, 25),
        datetime.date(2026, 3, 26),
        datetime.date(2026, 3, 27),
        datetime.date(2026, 3, 28),
        datetime.date(2026, 3, 29),
        datetime.date(2026, 3, 30),
        datetime.date(2026, 3, 31),
    ]
    sideridou_unavailability = [
        datetime.date(2026, 3, 16),
        datetime.date(2026, 3, 17),
        datetime.date(2026, 3, 18),
        datetime.date(2026, 3, 19),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
        datetime.date(2026, 3, 23),
        datetime.date(2026, 3, 24),
        datetime.date(2026, 3, 25),
    ]
    zafeiratou_unavailability = [
        datetime.date(2026, 3, 7),
        datetime.date(2026, 3, 25),
    ]
    kappa_unavailability = [
        datetime.date(2026, 3, 6),
        datetime.date(2026, 3, 11),
        datetime.date(2026, 3, 13),
        datetime.date(2026, 3, 18),
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 27),
    ]
    theofanopoulos_unavailability = [
        datetime.date(2026, 3, 20),
        datetime.date(2026, 3, 21),
        datetime.date(2026, 3, 22),
    ]

    test_shifts = {
        "ER1": Shift(name="ER1", doctors_per_shift=5, grants_day_off=False),
        "ER2": Shift(name="ER2", doctors_per_shift=4, grants_day_off=True),
        "Orofos": Shift(name="Orofos", doctors_per_shift=2, grants_day_off=True),
        "Eisagogeas": Shift(name="Eisagogeas", doctors_per_shift=1, grants_day_off=True),
        "Voitheia": Shift(name="Voitheia", doctors_per_shift=1, grants_day_off=False),
    }

    erotokritou_pre_assignemnts = [
        (datetime.date(2026, 3, 2), test_shifts["ER1"]),
        (datetime.date(2026, 3, 4), test_shifts["ER1"]),
        (datetime.date(2026, 3, 6), test_shifts["ER1"]),
        (datetime.date(2026, 3, 9), test_shifts["ER1"]),
        (datetime.date(2026, 3, 11), test_shifts["ER1"]),
    ]

    genikos1_pre_assignments = [
        (datetime.date(2026, 3, 2), test_shifts["ER1"]),
        (datetime.date(2026, 3, 4), test_shifts["ER1"]),
        (datetime.date(2026, 3, 6), test_shifts["ER1"]),
        (datetime.date(2026, 3, 9), test_shifts["ER1"]),
        (datetime.date(2026, 3, 11), test_shifts["ER1"]),
        (datetime.date(2026, 3, 13), test_shifts["ER1"]),
        (datetime.date(2026, 3, 16), test_shifts["ER1"]),
        (datetime.date(2026, 3, 18), test_shifts["ER1"]),
        (datetime.date(2026, 3, 20), test_shifts["ER1"]),
        (datetime.date(2026, 3, 23), test_shifts["ER1"]),
        (datetime.date(2026, 3, 27), test_shifts["ER1"]),
        (datetime.date(2026, 3, 30), test_shifts["ER1"]),
    ]

    genikos2_pre_assignments = [
        (datetime.date(2026, 3, 2), test_shifts["ER2"]),
        (datetime.date(2026, 3, 4), test_shifts["ER2"]),
        (datetime.date(2026, 3, 6), test_shifts["ER2"]),
        (datetime.date(2026, 3, 9), test_shifts["ER2"]),
        (datetime.date(2026, 3, 11), test_shifts["ER2"]),
        (datetime.date(2026, 3, 13), test_shifts["ER2"]),
        (datetime.date(2026, 3, 16), test_shifts["ER2"]),
        (datetime.date(2026, 3, 18), test_shifts["ER2"]),
        (datetime.date(2026, 3, 20), test_shifts["ER2"]),
        (datetime.date(2026, 3, 23), test_shifts["ER2"]),
        (datetime.date(2026, 3, 25), test_shifts["ER2"]),
        (datetime.date(2026, 3, 27), test_shifts["ER2"]),
        (datetime.date(2026, 3, 30), test_shifts["ER2"]),
        (datetime.date(2026, 3, 7), test_shifts["ER2"]),
        (datetime.date(2026, 3, 14), test_shifts["ER2"]),
        (datetime.date(2026, 3, 21), test_shifts["ER2"]),
        (datetime.date(2026, 3, 28), test_shifts["ER2"]),
    ]

    # genikos3_pre_assignments = [
    # ]

    doctors = [
        Doctor(name="Polyzou", email="polyzou@test.com",
               unavailability=set(polyzou_unavailability)),
        Doctor(name="Eleftherakis", email="elef@test.com",
               unavailability=eleftherakis_unavailability),
        Doctor(name="Michailidis", email="mich@test.com",
               unavailability=michailidis_unavailability),
        Doctor(name="Petropoulou", email="petrop@test.com",
               unavailability=petropoulou_unavailability),
        Doctor(name="Kassara", email="kassara@test.com",
               unavailability=kassara_unavailability),
        Doctor(name="Marlafeka", email="marla@test.com",
               unavailability=marlafeka_unavailability),
        Doctor(name="Efthymiou", email="efth@test.com",
               unavailability=efthymiou_unavailability),
        Doctor(name="Triantafyllopoulos", email="triant@test.com",
               unavailability=triantafyllopoulos_unavailability),
        Doctor(name="Papakamenos", email="papakam@test.com",
               unavailability=papakamenos_unavailability),
        Doctor(name="Karydis", email="karydis@test.com",
               unavailability=karydis_unavailability),
        Doctor(name="Aravantinos", email="arav@test.com",
               unavailability=aravantinos_unavailability),
        Doctor(name="Argyropoulos", email="argy@test.com",
               unavailability=argyropoulos_unavailability),
        Doctor(name="Xiotis", email="xiotis@test.com",
               unavailability=xiotis_unavailability),
        Doctor(name="Nikas", email="nikas@test.com",
               unavailability=nikas_unavailability),
        Doctor(name="Adamopoulou", email="adam@test.com",
               unavailability=adamopoulou_unavailability),
        Doctor(name="Mpakalarou", email="mpak@test.com",
               unavailability=mpakalarou_unavailability),
        Doctor(name="Papageorgiou Despoina", email="papagdsp@test.com",
               unavailability=papageorgiou_dsp_unavailability),
        Doctor(name="Rousia", email="rousia@test.com",
               unavailability=rousia_unavailability),
        Doctor(name="Pitsi", email="pitsi@test.com",
               unavailability=pitsi_unavailability),
        Doctor(name="Moutsi", email="moutsi@test.com",
               unavailability=moutsi_unavailability),
        Doctor(name="Tzikopoulos", email="tzik@test.com",
               unavailability=tzikopoulos_unavailability),
        Doctor(name="Giannaki", email="giann@test.com",
               unavailability=giannaki_unavailability),
        Doctor(name="Gkolemi", email="gkol@test.com",
               unavailability=gkolemi_unavailability),
        Doctor(name="Tasiopoulos", email="tasiop@test.com",
               unavailability=tasiopoulos_unavailability),
        Doctor(name="Florou", email="florou@test.com",
               unavailability=florou_unavailability),
        Doctor(name="Spiliotopoulos", email="florou@test.com",
               unavailability=spiliotopoulos_unavailability),
        Doctor(name="Xrysanthakopoulou", email="xrys@test.com",
               unavailability=xrysanthakopoulos_unavailability),
        Doctor(name="Pandi", email="pandi@test.com",
               unavailability=pandi_unavailability),
        Doctor(name="Argyriadi", email="argyriadi@test.com",
               unavailability=argyriadi_unavailability),
        Doctor(name="Armeni", email="arm@test.com",
               unavailability=armeni_unavailability),
        Doctor(name="Papadopoulos", email="papadop@test.com",
               unavailability=papadopoulos_unavailability),
        Doctor(name="Erotokritou", email="eroto@test.com",
               unavailability=erotokritou_unavailability,
               pre_assignments=erotokritou_pre_assignemnts
               ),
        Doctor(name="Sideratou", email="sidera@test.com",
               unavailability=sideridou_unavailability),
        Doctor(name="Zafeiratou", email="zaf@test.com",
               unavailability=zafeiratou_unavailability),
        Doctor(name="Kapp", email="kappa@test.com",
               unavailability=kappa_unavailability),
        Doctor(name="Theofanopoulos", email="theof@test.com",
               unavailability=theofanopoulos_unavailability),
        Doctor(name="Genikos1", email="gen1@test.com", unavailability=[],
               pre_assignments=genikos1_pre_assignments),
        Doctor(name="Genikos2", email="gen2@test.com",
               unavailability=[], pre_assignments=genikos2_pre_assignments),
        # Doctor(name="Genikos3", email="gen3@test.com",
        #        unavailability=[], pre_assignments=genikos3_pre_assignments)
    ]

    positions = [
        Position(name="Orofos", shifts=[test_shifts["Orofos"]], eligible_doctors=[
                 doctors[14], *doctors[16:26]]),
        Position(name="Eisagogeas",
                 shifts=[test_shifts["Eisagogeas"]], eligible_doctors=doctors[2:9], duty_days=set(
                     [0, 2, 4, 5])),
        Position(name="Voitheia",
                 shifts=[test_shifts["Voitheia"]], eligible_doctors=doctors[2:9], duty_days={6}),
        Position(name="ER", shifts=[test_shifts["ER1"], test_shifts["ER2"]], duty_days=set(
            [0, 2, 4, 5]), eligible_doctors=[*doctors[0:16], *doctors[26:]]),
    ]

    teams = [
        Team(name="Team A1", doctors=[doctors[7], doctors[16], doctors[18]]),
        Team(name="Team A2", doctors=[doctors[10],
             doctors[24], doctors[25], doctors[23]]),
        Team(name="Team B1", doctors=[doctors[3], doctors[13], doctors[17]]),
        Team(name="Team B2", doctors=[doctors[11], doctors[22], doctors[12]]),
        Team(name="Team D", doctors=[
             doctors[6], doctors[5], doctors[9], doctors[14], doctors[15], doctors[20], doctors[22]
             ]),
    ]

    # department = Department(name="Pathologia", positions=positions, teams=teams, teamless_doctors=[
    #                         *doctors[0:3], doctors[4], doctors[8], doctors[19], doctors[21], *doctors[26:]])

    test_department = Department(
        name="Pathologia",
        positions=positions,
        teams=teams,
        doctor_order=doctors,
        teamless_doctors=[
            *doctors[0:3], doctors[4], doctors[8], doctors[19], doctors[21], *doctors[26:]
        ]
    )

    app = ShiftScheduler(department=test_department)
    app.create_schedule(month=3, year=2026)
    print(app.dates[0], "→", app.dates[-1])
    print("Days in month:", len(app.dates))
    print("Weekend days:", sum(
        1 for d in app.dates if app.is_weekend[d]))

    print("\n\n")
    print("--------Schedule---------\n")
    app.print_schedule()

    app.export_to_exel("schedule.xlsx")
