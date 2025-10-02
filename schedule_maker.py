import pandas as pd
import datetime as dt
from ortools.sat.python import cp_model
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import math

# ------------------------------
# User parameters / tweakable weights
# ------------------------------
INPUT_FILE = "input.xlsx"
OUT_FILE = "monthly_schedule.xlsx"

# weights for the combined objective (tweak to taste)
W_EVERY_OTHER_PENALTY = 4       # penalty for every-other patterns
W_GAP_PENALTY = 2               # penalty for short gaps (i.e., i and i+2)
W_BLOCK_DEV_PENALTY = 2         # penalty per unit deviation from ideal per block
W_FULL_WKEND_OFF_BONUS = 5      # reward for full weekend off (Fri+Sat+Sun)
W_BALANCE_FULL_WKENDS_OFF = 20  # weight for balancing weekend duties
W_DIFF_WKEND_DUTY_DAY = 2       # weight for balancing weekend days of duty

# solver time limit seconds
SOLVER_TIME_LIMIT = 120

# ------------------------------
# 1. Read input Excel
# ------------------------------
doctors_df = pd.read_excel(INPUT_FILE, sheet_name="Doctors")

# Choose month/year: default = next month
today = dt.date.today()
next_month_first = (today.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
month_for_schedule = next_month_first.month
year_for_schedule = next_month_first.year

# Option: you can override above two lines manually if you want a specific month:
# month_for_schedule = 12
# year_for_schedule = 2025

first_day = dt.date(year_for_schedule, month_for_schedule, 1)
import calendar
days_in_month = calendar.monthrange(year_for_schedule, month_for_schedule)[1]
last_day = dt.date(year_for_schedule, month_for_schedule, days_in_month)
dates = [first_day + dt.timedelta(days=i) for i in range((last_day - first_day).days + 1)]
print(f"Creating schedule for: {first_day} → {last_day}")

# Identify weekends
is_weekend = {d: (d.weekday() >= 5) for d in dates}

# Parse unavailability column (expects comma-separated day numbers, e.g. "1,2,15")
unavailability = {}
for idx, row in doctors_df.iterrows():
    doc = row["Doctor"]
    raw = row.get("Unavailability", "")
    if pd.isna(raw) or str(raw).strip() == "":
        unavailable_days = []
    else:
        # allow entries like "1, 2, 15" or "5"
        tokens = [t.strip() for t in str(raw).split(",") if t.strip() != ""]
        unavailable_days = []
        for t in tokens:
            try:
                day_num = int(t)
                if 1 <= day_num <= len(dates):
                    unavailable_days.append(dt.date(year_for_schedule, month_for_schedule, day_num))
                else:
                    print(f"Warning: day {day_num} out of range for {month_for_schedule}/{year_for_schedule}")
            except ValueError:
                # ignore bad tokens
                print(f"Warning: could not parse unavailability token '{t}' for doctor {doc}")
    doctors_df.at[idx, "Unavailability"] = unavailable_days
    unavailability[doc] = set(unavailable_days)

doctors_list = doctors_df["Doctor"].tolist()
num_docs = len(doctors_list)

# ------------------------------
# 2. Build the model
# ------------------------------
model = cp_model.CpModel()

# Create variables x[(i,doc)] = 1 if doc assigned on date index i
x = {}
for i, day in enumerate(dates):
    for doc in doctors_list:
        if day in unavailability.get(doc, set()):
            continue
        x[(i, doc)] = model.NewBoolVar(f"x_{i}_{doc}")

# Check if we have enough doctors to cover all days
max_possible_duties = len(doctors_list) * 7
allow_unassigned_tuesdays = False
if max_possible_duties < len(dates):
    print("⚠ Not enough doctors to cover all days. Tuesdays will be left unassigned.")
    allow_unassigned_tuesdays = True

# Exactly one doctor per day
for i, day in enumerate(dates):
    if allow_unassigned_tuesdays and day.weekday() == 1:  # 1 = Tuesday
        # Either 0 or 1 doctor (so it can be left unassigned)
        model.Add(sum(x.get((i,doc), 0) for doc in doctors_list) <= 1)
    else:
        # Normal rule: exactly 1 doctor assigned
        model.Add(sum(x.get((i,doc), 0) for doc in doctors_list) == 1)


# Balanced total duties per doctor (difference at most 1)
total_days = len(dates)
min_days = total_days // num_docs
max_days = min_days if total_days % num_docs == 0 else min_days + 1
for doc in doctors_list:
    model.Add(sum(x.get((i, doc), 0) for i in range(len(dates))) >= min_days)
    model.Add(sum(x.get((i, doc), 0) for i in range(len(dates))) <= max_days)
    
    # hard constraint: no more than 7 duties in the month
    model.Add(sum(x.get((i, doc), 0) for i in range(len(dates))) <= 7)

# No consecutive duties (hard constraint)
for i in range(len(dates) - 1):
    for doc in doctors_list:
        model.Add(x.get((i, doc), 0) + x.get((i + 1, doc), 0) <= 1)

# Balance number of weekend/weekday duties (difference at most 1)
total_weekends = sum(1 for d in dates if is_weekend[d])
min_wkend = total_weekends // num_docs
max_wkend = min_wkend if total_weekends % num_docs == 0 else min_wkend + 1
for doc in doctors_list:
    model.Add(sum(x.get((i, doc), 0) for i, d in enumerate(dates) if is_weekend[d]) >= min_wkend)
    model.Add(sum(x.get((i, doc), 0) for i, d in enumerate(dates) if is_weekend[d]) <= max_wkend)


# ------------------------------
# Soft preference variables (we'll combine into one objective)
# ------------------------------
every_other_vars = []     # patterns i and i+2 both assigned
gap2_vars = []            # same as every_other, kept separately if you want different weights
block_deviation_vars = []    # deviation from ideal per block
full_weekend_off_bonus = []  # reward for full weekend off (Fri+Sat+Sun)
balanced_full_wkends_off_deviation_vars = [] # deviation vars for balancing full weekends off
different_weekend_duty_day_vars = []  # vars to balance weekend duty days

# penalize every-other patterns (i,i+2)
for doc in doctors_list:
    for i in range(len(dates) - 2):
        if (i, doc) in x and (i + 2, doc) in x:
            b = model.NewBoolVar(f"everyother_{i}_{doc}")
            model.Add(x[(i, doc)] + x[(i + 2, doc)] == 2).OnlyEnforceIf(b)
            model.Add(x[(i, doc)] + x[(i + 2, doc)] != 2).OnlyEnforceIf(b.Not())
            every_other_vars.append(b)
            gap2_vars.append(b)

# Block balancing: split month into blocks (4 blocks) and penalize deviation from ideal per block
num_blocks = 4
block_size = math.ceil(len(dates) / num_blocks)
ideal_total_per_doc = total_days / num_docs
ideal_per_block = ideal_total_per_doc / num_blocks  # target duties per doc per block (float)
# We'll create integer deviation vars capturing absolute deviation from rounded ideal
for doc in doctors_list:
    for b in range(num_blocks):
        start = b * block_size
        end = min((b + 1) * block_size, len(dates))
        if start >= end:
            continue
        duties_vars = [x[(i, doc)] for i in range(start, end) if (i, doc) in x]
        if not duties_vars:
            # doc unavailable for entire block - create a 0 deviation (no var)
            continue
        duties_sum = sum(duties_vars)
        # rounded ideal for block (we want integer target near ideal)
        rounded_ideal_low = int(math.floor(ideal_per_block))
        rounded_ideal_high = int(math.ceil(ideal_per_block))
        # We'll allow deviation from rounded ideal; create a deviation variable "dev >= abs(duties_sum - rounded_ideal)"
        dev = model.NewIntVar(0, len(dates), f"dev_block_{b}_{doc}")
        # dev >= duties_sum - rounded_ideal_high
        model.Add(duties_sum - rounded_ideal_high <= dev)
        # dev >= rounded_ideal_low - duties_sum
        model.Add(rounded_ideal_low - duties_sum <= dev)
        block_deviation_vars.append(dev)
      

# Count full weekends off per doctor
weekends_off_per_doc = {
    doc: [] for doc in doctors_list
}      
        
# Full weekend off bonus (Fri+Sat+Sun)
# Reward if a doctor has an entire weekend off (Fri, Sat, Sun)
for i, day in enumerate(dates):
    if day.weekday() == 4:  # Friday
        sat_idx = i + 1 if i + 1 < len(dates) else None
        sun_idx = i + 2 if i + 2 < len(dates) else None
        if sat_idx is not None and sun_idx is not None:
            for doc in doctors_list:
                vars_window = [
                    x.get((i,doc), 0),
                    x.get((sat_idx,doc), 0),
                    x.get((sun_idx,doc), 0)
                ]
                b = model.NewBoolVar(f"full_wkend_off_{i}_{doc}")
                model.Add(sum(vars_window) == 0).OnlyEnforceIf(b)
                model.Add(sum(vars_window) != 0).OnlyEnforceIf(b.Not())
                full_weekend_off_bonus.append(b)
                weekends_off_per_doc[doc].append(b)

total_full_weekends = len([d for d in dates if d.weekday() == 4])  # number of Fridays
min_wkend_off = total_full_weekends // len(doctors_list)
max_wkend_off = min_wkend_off if total_full_weekends % len(doctors_list) == 0 else min_wkend_off + 1

full_weekends_off_count = {
    doc: sum(weekends_off_per_doc[doc]) for doc in doctors_list
}
    
avg_full_weekends_off = total_full_weekends / len(doctors_list)

for doc in doctors_list:
    diff = model.NewIntVar(0, total_full_weekends, f"diff_{doc}")
    model.Add(diff >= full_weekends_off_count[doc] - int(avg_full_weekends_off))
    model.Add(diff >= int(avg_full_weekends_off) - full_weekends_off_count[doc])
    balanced_full_wkends_off_deviation_vars.append(diff)

# Collect weekend indices
saturdays = [i for i, d in enumerate(dates) if d.weekday() == 5]
sundays = [i for i, d in enumerate(dates) if d.weekday() == 6]

for doc in doctors_list:
    # Saturdays
    sat_vars = [x[(i,doc)] for i in saturdays if (i,doc) in x]
    if len(sat_vars) > 1:
        extra_sat = model.NewIntVar(0, len(sat_vars)-1, f"extra_sat_{doc}")
        model.Add(extra_sat == sum(sat_vars) - 1)
        different_weekend_duty_day_vars.append(extra_sat)
    # Sundays
    sun_vars = [x[(i,doc)] for i in sundays if (i,doc) in x]
    if len(sun_vars) > 1:
        extra_sun = model.NewIntVar(0, len(sun_vars)-1, f"extra_sun_{doc}")
        model.Add(extra_sun == sum(sun_vars) - 1)
        different_weekend_duty_day_vars.append(extra_sun)

for doc in doctors_list:
    # Collect all "full weekend" indices: Fri, Sat, Sun sequences
    full_weekends = []
    for i in range(len(dates) - 2):
        if dates[i].weekday() == 4 and dates[i+1].weekday() == 5 and dates[i+2].weekday() == 6:
            full_weekends.append((i, i+1, i+2))

    if full_weekends:
        # For each full weekend, create a bool var that is 1 if doctor is OFF all three days
        off_weekend_vars = []
        for fri, sat, sun in full_weekends:
            off_var = model.NewBoolVar(f"off_full_weekend_{doc}_{fri}")
            
            # Check if doctor is assigned on each day (None if unavailable)
            fri_assigned = x.get((fri, doc), None)
            sat_assigned = x.get((sat, doc), None)
            sun_assigned = x.get((sun, doc), None)
            
            # Treat unavailable days as already off (0)
            fri_val = fri_assigned if fri_assigned is not None else model.NewConstant(0)
            sat_val = sat_assigned if sat_assigned is not None else model.NewConstant(0)
            sun_val = sun_assigned if sun_assigned is not None else model.NewConstant(0)
            
            # off_var = 1 if all three days are free
            model.Add(fri_val + sat_val + sun_val == 0).OnlyEnforceIf(off_var)
            model.Add(fri_val + sat_val + sun_val != 0).OnlyEnforceIf(off_var.Not())
            off_weekend_vars.append(off_var)
        
        # Hard constraint: at least one full weekend off
        model.Add(sum(off_weekend_vars) >= 1)


# ------------------------------
# Combine objective into one expression
# ------------------------------
# We want to MAXIMIZE good things and MINIMIZE bad things.
# Convert penalties to negatives inside the Maximize expression.

obj_terms = []

# Full weekend off rewards (positive)
if full_weekend_off_bonus:
    obj_terms.append(W_FULL_WKEND_OFF_BONUS * sum(full_weekend_off_bonus))

# Penalties (negative)
# if every_other_vars:
#     obj_terms.append(-W_EVERY_OTHER_PENALTY * sum(every_other_vars))

# if gap2_vars:
#     obj_terms.append(-W_GAP_PENALTY * sum(gap2_vars))

if block_deviation_vars:
    # block_deviation_vars are IntVars — penalize sum of deviations
    obj_terms.append(-W_BLOCK_DEV_PENALTY * sum(block_deviation_vars))
    
if balanced_full_wkends_off_deviation_vars:
    obj_terms.append(-W_BALANCE_FULL_WKENDS_OFF * sum(balanced_full_wkends_off_deviation_vars))
    
if different_weekend_duty_day_vars:
    obj_terms.append(-W_DIFF_WKEND_DUTY_DAY * sum(different_weekend_duty_day_vars))

# If no soft terms, minimize nothing (but that shouldn't be the case)
full_obj = sum(obj_terms) if obj_terms else 0
model.Maximize(full_obj)

# ------------------------------
# 3. Solve
# ------------------------------
solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = SOLVER_TIME_LIMIT
solver.parameters.num_search_workers = 8
status = solver.Solve(model)

if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    raise RuntimeError("No feasible schedule found.")

# ------------------------------
# 4. Export to Excel
# ------------------------------
schedule = []
for i, day in enumerate(dates):
    assigned = None
    for doc in doctors_list:
        var = x.get((i, doc), None)
        if var is not None and solver.Value(var) == 1:
            assigned = doc
            break
    if assigned is None:
        # should not happen (we forced 1 per day), but be safe
        assigned = "UNASSIGNED"
    schedule.append({"Date": day, "Assigned Doctor": assigned})

schedule_df = pd.DataFrame(schedule)
schedule_df.to_excel(OUT_FILE, index=False)
print(f"Schedule created: {OUT_FILE}")

# --- Apply styling with openpyxl ---
wb = load_workbook(OUT_FILE)
ws = wb.active

weekend_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
weekend_font = Font(color="FFFFFF", bold=True)

for row in range(2, ws.max_row + 1):  # skip header
    date_cell = ws.cell(row=row, column=1)
    date_value = pd.to_datetime(date_cell.value).date()
    if date_value.weekday() >= 5:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = weekend_fill
            cell.font = weekend_font

# Adjust column width based on max length in each column
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    # Add some padding (e.g., +2)
    adjusted_width = max_length + 2
    ws.column_dimensions[col_letter].width = adjusted_width

wb.save(OUT_FILE)

# ------------------------------
# Diagnostics: print counts
# ------------------------------
print("\nDuties per doctor:")
for doc in doctors_list:
    duties = 0
    for i in range(len(dates)):
        var = x.get((i, doc), None)
        if var is not None and solver.Value(var) == 1:
            duties += 1
    print(f"  {doc}: {duties}")

print("\nWeekend duties per doctor:")
for doc in doctors_list:
    wend = 0
    for i, d in enumerate(dates):
        if is_weekend[d]:
            var = x.get((i, doc), None)
            if var is not None and solver.Value(var) == 1:
                wend += 1
    print(f"  {doc}: {wend}")

print("\nAssigned days per doctor (day num + weekday):")
for doc in doctors_list:
    assigned_days = []
    for i in range(len(dates)):
        var = x.get((i, doc), None)
        if var is not None and solver.Value(var) == 1:
            assigned_days.append(dates[i].strftime("%d %a"))
    print(f"  {doc}: {', '.join(assigned_days)}")
    
print("\nFull weekends off (Fri+Sat+Sun) per doctor:")
for doc in doctors_list:
    full_wkends = 0
    for i, day in enumerate(dates):
        if day.weekday() == 4:  # Friday
            sat_idx = i + 1 if i + 1 < len(dates) else None
            sun_idx = i + 2 if i + 2 < len(dates) else None
            if sat_idx is not None and sun_idx is not None:
                vars_window = [
                    x.get((i,doc), 0),        # Friday
                    x.get((sat_idx,doc), 0),  # Saturday
                    x.get((sun_idx,doc), 0)   # Sunday
                ]
                if all(solver.Value(v) == 0 for v in vars_window):
                    full_wkends += 1
    print(f"  {doc}: {full_wkends}")
